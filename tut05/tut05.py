import csv
import os
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

if not os.path.exists("output"):
    os.makedirs("output")

with open("names-roll.csv", 'r') as file:
   rows=csv.reader(file)
   rlnm={} 
   for line in rows:
      if(line[0]=="Roll"):continue
      rlnm[line[0]]=f"{line[1].strip()}"

with open("subjects_master.csv" , 'r') as file:
    rows=csv.reader(file)
    sbj={}
    for line in rows:
        if line[0]== 'subno':continue
        sbj[line[0]]=[line[1],line[2]]

with open("grades.csv",'r') as file:
    rows = csv.reader(file)
    for line in rows:
        if line[0] == 'Roll':continue
        with open(f"output\\{line[0]}_{line[1]}.csv" , "a" , newline='') as file:
            writer=csv.writer(file)
            writer.writerow([line[2],sbj[line[2]][0] ,sbj[line[2]][1] ,line[3] ,line[5] , line[4]])

rplc={"AA":"10","AB":"9","BB":"8","BC":"7","CC":"6","CD":"5","DD":"4","F":"0","I":"0","F*":"0","I*":"0","DD*":"4"," BB":"8"}
s1c=  { x:0 for x in rlnm}
s2c=  { x:0 for x in rlnm}
s3c=  { x:0 for x in rlnm}
s4c=  { x:0 for x in rlnm}
s5c=  { x:0 for x in rlnm}
s6c=  { x:0 for x in rlnm}
s7c=  { x:0 for x in rlnm}
s8c=  { x:0 for x in rlnm}
s10c=  { x:0 for x in rlnm}
s1u=  { x:0 for x in rlnm}
s2u=  { x:0 for x in rlnm}
s3u=  { x:0 for x in rlnm}
s4u=  { x:0 for x in rlnm}
s5u=  { x:0 for x in rlnm}
s6u=  { x:0 for x in rlnm}
s7u=  { x:0 for x in rlnm}
s8u=  { x:0 for x in rlnm}
s10u=  { x:0 for x in rlnm}


for key,value in rlnm.items():
    wb = Workbook()    
    wb.remove(wb["Sheet"])            
    if(os.path.exists(f"output\\{key}_1.csv")) :
      wb.create_sheet(index=1, title="Sem1")  
      Sem1=wb["Sem1"]
      Sem1.append(["Sl No.","Subject No.","Subject Name","L-T-P","Credit","Subject Type","Grade"])
      with open(f"output\\{key}_1.csv", 'r') as file:
        rows=csv.reader(file)
        for line in rows:
            s1c[key]=int(s1c[key])+int(line[3])
            s1u[key]=int(s1u[key])+int(line[3])*int(rplc[line[5]])
            row_count = Sem1.max_row
            line.insert(0, row_count)           
            Sem1.append(line)
      os.remove(f"output\\{key}_1.csv")      
    if(os.path.exists(f"output\\{key}_2.csv")) :
      wb.create_sheet(index=2, title="Sem2")  
      Sem2=wb["Sem2"]
      Sem2.append(["Sl No.","Subject No.","Subject Name","L-T-P","Credit","Subject Type","Grade"])
      for row in Sem2.iter_rows(min_row=1, max_row=1):
          for cell in row:
              cell.font = Font(bold=True)
              cell.fill = PatternFill(fgColor="77C3FD", fill_type = "solid")
      with open(f"output\\{key}_2.csv", 'r') as file:
        rows=csv.reader(file)
        for line in rows:
            s2c[key]=int(s2c[key])+int(line[3])
            s2u[key]=int(s2u[key])+int(line[3])*int(rplc[line[5]])
            row_count = Sem2.max_row
            line.insert(0, row_count)           
            Sem2.append(line)
        for row in Sem2.iter_rows(min_row=2,max_row=row_count+1):
           for cell in row:
              cell.font = Font(bold=True)
              cell.fill = PatternFill(fgColor="B5DDFB", fill_type = "solid")         
      os.remove(f"output\\{key}_2.csv")                    
    if(os.path.exists(f"output\\{key}_3.csv")) :
      wb.create_sheet(index=3, title="Sem3")  
      Sem3=wb["Sem3"]
      Sem3.append(["Sl No.","Subject No.","Subject Name","L-T-P","Credit","Subject Type","Grade"])
      with open(f"output\\{key}_3.csv", 'r') as file:
        rows=csv.reader(file)
        for line in rows:
            s3c[key]=int(s3c[key])+int(line[3])
            s3u[key]=int(s3u[key])+int(line[3])*int(rplc[line[5]])
            row_count = Sem3.max_row
            line.insert(0, row_count)           
            Sem3.append(line)
      os.remove(f"output\\{key}_3.csv")        
    if(os.path.exists(f"output\\{key}_4.csv")) :
      wb.create_sheet(index=4, title="Sem4")  
      Sem4=wb["Sem4"]
      Sem4.append(["Sl No.","Subject No.","Subject Name","L-T-P","Credit","Subject Type","Grade"])
      with open(f"output\\{key}_4.csv", 'r') as file:
        rows=csv.reader(file)
        for line in rows:
            s4c[key]=int(s4c[key])+int(line[3])
            s4u[key]=int(s4u[key])+int(line[3])*int(rplc[line[5]])
            row_count = Sem4.max_row
            line.insert(0, row_count)           
            Sem4.append(line)
      os.remove(f"output\\{key}_4.csv")        
    if(os.path.exists(f"output\\{key}_5.csv")) :
      wb.create_sheet(index=5, title="Sem5")  
      Sem5=wb["Sem5"]
      Sem5.append(["Sl No.","Subject No.","Subject Name","L-T-P","Credit","Subject Type","Grade"])
      with open(f"output\\{key}_5.csv", 'r') as file:
        rows=csv.reader(file)
        for line in rows:
            s5c[key]=int(s5c[key])+int(line[3])
            s5u[key]=int(s5u[key])+int(line[3])*int(rplc[line[5]])
            row_count = Sem5.max_row
            line.insert(0, row_count)           
            Sem5.append(line)  
      os.remove(f"output\\{key}_5.csv")                             
    if(os.path.exists(f"output\\{key}_6.csv")) :
      wb.create_sheet(index=6, title="Sem6")  
      Sem6=wb["Sem6"]
      Sem6.append(["Sl No.","Subject No.","Subject Name","L-T-P","Credit","Subject Type","Grade"])
      with open(f"output\\{key}_6.csv", 'r') as file:
        rows=csv.reader(file)
        for line in rows:
            s6c[key]=int(s6c[key])+int(line[3])
            s6u[key]=int(s6u[key])+int(line[3])*int(rplc[line[5]])
            row_count = Sem6.max_row
            line.insert(0, row_count)           
            Sem6.append(line)           
      os.remove(f"output\\{key}_6.csv")        
    if(os.path.exists(f"output\\{key}_7.csv")) :
      wb.create_sheet(index=7, title="Sem7")  
      Sem7=wb["Sem7"]
      Sem7.append(["Sl No.","Subject No.","Subject Name","L-T-P","Credit","Subject Type","Grade"])
      with open(f"output\\{key}_7.csv", 'r') as file:
        rows=csv.reader(file)
        for line in rows:
            s7c[key]=int(s7c[key])+int(line[3])
            s7u[key]=int(s7u[key])+int(line[3])*int(rplc[line[5]])
            row_count = Sem7.max_row
            line.insert(0, row_count)           
            Sem7.append(line)  
      os.remove(f"output\\{key}_7.csv")                  
    if(os.path.exists(f"output\\{key}_8.csv")) :
      wb.create_sheet(index=8, title="Sem8")  
      Sem8=wb["Sem8"]
      Sem8.append(["Sl No.","Subject No.","Subject Name","L-T-P","Credit","Subject Type","Grade"])
      with open(f"output\\{key}_8.csv", 'r') as file:
        rows=csv.reader(file)
        for line in rows:
            s8c[key]=int(s8c[key])+int(line[3])
            s8u[key]=int(s8u[key])+int(line[3])*int(rplc[line[5]])
            row_count = Sem8.max_row
            line.insert(0, row_count)           
            Sem8.append(line)
      os.remove(f"output\\{key}_8.csv")        
    if(os.path.exists(f"output\\{key}_10.csv")) :
      wb.create_sheet(index=10, title="Sem10")  
      Sem10=wb["Sem10"]
      Sem10.append(["Sl No.","Subject No.","Subject Name","L-T-P","Credit","Subject Type","Grade"])
      with open(f"output\\{key}_10.csv", 'r') as file:
        rows=csv.reader(file)
        for line in rows:
            s10c[key]=int(s10c[key])+int(line[3])
            s10u[key]=int(s10u[key])+int(line[3])*int(rplc[line[5]])
            row_count = Sem10.max_row
            line.insert(0, row_count)           
            Sem10.append(line)
      os.remove(f"output\\{key}_10.csv")        

    wb.create_sheet(index=0, title="Overall")  
    Overall=wb["Overall"]
    if(s3c[key]==0):
       Overall.append(["Roll No.",f"{key}"])
       Overall.append(["Name of Student",f"{value}"])
       Overall.append(["Discipline",f"{key[4:6]}"])
       Overall.append(["Semester No.","1","2"])
       Overall.append(["Semester wise Credit Taken",f"{s1c[key]}",f"{s2c[key]}"])
       Overall.append(["SPI","{:.2f}".format(s1u[key]/s1c[key]),"{:.2f}".format(s2u[key]/s2c[key])])
       Overall.append(["Total Credits Taken",f"{s1c[key]}",f"{s1c[key]+s2c[key]}"])
       Overall.append(["CPI","{:.2f}".format(s1u[key]/s1c[key]),"{:.2f}".format((s1u[key]+s2u[key])/(s1c[key]+s2c[key]))])
    elif(s5c[key]==0):
        Overall.append(["Roll No.",f"{key}"])
        Overall.append(["Name of Student",f"{value}"])
        Overall.append(["Discipline",f"{key[4:6]}"])
        Overall.append(["Semester No.","1","2","3","4"])
        Overall.append(["Semester wise Credit Taken",f"{s1c[key]}",f"{s2c[key]}",f"{s3c[key]}",f"{s4c[key]}"])
        Overall.append(["SPI","{:.2f}".format(s1u[key]/s1c[key]),"{:.2f}".format(s2u[key]/s2c[key]),"{:.2f}".format(s3u[key]/s3c[key]),"{:.2f}".format(s4u[key]/s4c[key])])
        Overall.append(["Total Credits Taken",f"{s1c[key]}",f"{s1c[key]+s2c[key]}",f"{s1c[key]+s2c[key]+s3c[key]}",f"{s1c[key]+s2c[key]+s3c[key]+s4c[key]}"])
        Overall.append(["CPI","{:.2f}".format(s1u[key]/s1c[key]),"{:.2f}".format((s1u[key]+s2u[key])/(s1c[key]+s2c[key])),"{:.2f}".format((s1u[key]+s2u[key]+s3u[key])/(s1c[key]+s2c[key]+s3c[key])),"{:.2f}".format((s1u[key]+s2u[key]+s3u[key]+s4u[key])/(s1c[key]+s2c[key]+s3c[key]+s4c[key]))])
    elif(s6c[key]==0): 
        Overall.append(["Roll No.",f"{key}"])
        Overall.append(["Name of Student",f"{value}"])
        Overall.append(["Discipline",f"{key[4:6]}"])
        Overall.append(["Semester No.","1","2","3","4","5"])
        Overall.append(["Semester wise Credit Taken",f"{s1c[key]}",f"{s2c[key]}",f"{s3c[key]}",f"{s4c[key]}",f"{s5c[key]}"])
        Overall.append(["SPI","{:.2f}".format(s1u[key]/s1c[key]),"{:.2f}".format(s2u[key]/s2c[key]),"{:.2f}".format(s3u[key]/s3c[key]),"{:.2f}".format(s4u[key]/s4c[key]),"{:.2f}".format(s5u[key]/s5c[key])])
        Overall.append(["Total Credits Taken",f"{s1c[key]}",f"{s1c[key]+s2c[key]}",f"{s1c[key]+s2c[key]+s3c[key]}",f"{s1c[key]+s2c[key]+s3c[key]+s4c[key]}",f"{s1c[key]+s2c[key]+s3c[key]+s4c[key]+s5c[key]}"])
        Overall.append(["CPI","{:.2f}".format(s1u[key]/s1c[key]),"{:.2f}".format((s1u[key]+s2u[key])/(s1c[key]+s2c[key])),"{:.2f}".format((s1u[key]+s2u[key]+s3u[key])/(s1c[key]+s2c[key]+s3c[key])),"{:.2f}".format((s1u[key]+s2u[key]+s3u[key]+s4u[key])/(s1c[key]+s2c[key]+s3c[key]+s4c[key])),"{:.2f}".format((s1u[key]+s2u[key]+s3u[key]+s4u[key]+s5u[key])/(s1c[key]+s2c[key]+s3c[key]+s4c[key]+s5c[key]))])
    elif(s8c[key]==0):
        Overall.append(["Roll No.",f"{key}"])
        Overall.append(["Name of Student",f"{value}"])
        Overall.append(["Discipline",f"{key[4:6]}"])
        Overall.append(["Semester No.","1","2","3","4","5","6","7"])
        Overall.append(["Semester wise Credit Taken",f"{s1c[key]}",f"{s2c[key]}",f"{s3c[key]}",f"{s4c[key]}",f"{s5c[key]}",f"{s6c[key]}",f"{s7c[key]}"])
        Overall.append(["SPI","{:.2f}".format(s1u[key]/s1c[key]),"{:.2f}".format(s2u[key]/s2c[key]),"{:.2f}".format(s3u[key]/s3c[key]),"{:.2f}".format(s4u[key]/s4c[key]),"{:.2f}".format(s5u[key]/s5c[key]),"{:.2f}".format(s6u[key]/s6c[key]),"{:.2f}".format(s7u[key]/s7c[key])])
        Overall.append(["Total Credits Taken",f"{s1c[key]}",f"{s1c[key]+s2c[key]}",f"{s1c[key]+s2c[key]+s3c[key]}",f"{s1c[key]+s2c[key]+s3c[key]+s4c[key]}",f"{s1c[key]+s2c[key]+s3c[key]+s4c[key]+s5c[key]}",f"{s1c[key]+s2c[key]+s3c[key]+s4c[key]+s5c[key]+s6c[key]}",f"{s1c[key]+s2c[key]+s3c[key]+s4c[key]+s5c[key]+s6c[key]+s7c[key]}"])
        Overall.append(["CPI","{:.2f}".format(s1u[key]/s1c[key]),"{:.2f}".format((s1u[key]+s2u[key])/(s1c[key]+s2c[key])),"{:.2f}".format((s1u[key]+s2u[key]+s3u[key])/(s1c[key]+s2c[key]+s3c[key])),"{:.2f}".format((s1u[key]+s2u[key]+s3u[key]+s4u[key])/(s1c[key]+s2c[key]+s3c[key]+s4c[key])),"{:.2f}".format((s1u[key]+s2u[key]+s3u[key]+s4u[key]+s5u[key])/(s1c[key]+s2c[key]+s3c[key]+s4c[key]+s5c[key])),"{:.2f}".format((s1u[key]+s2u[key]+s3u[key]+s4u[key]+s5u[key]+s6u[key])/(s1c[key]+s2c[key]+s3c[key]+s4c[key]+s5c[key]+s6c[key])),"{:.2f}".format((s1u[key]+s2u[key]+s3u[key]+s4u[key]+s5u[key]+s6u[key]+s7u[key])/(s1c[key]+s2c[key]+s3c[key]+s4c[key]+s5c[key]+s6c[key]+s7c[key]))])
    elif(s10c[key]==0):
        Overall.append(["Roll No.",f"{key}"])
        Overall.append(["Name of Student",f"{value}"])
        Overall.append(["Discipline",f"{key[4:6]}"])
        Overall.append(["Semester No.","1","2","3","4","5","6","7","8"])
        Overall.append(["Semester wise Credit Taken",f"{s1c[key]}",f"{s2c[key]}",f"{s3c[key]}",f"{s4c[key]}",f"{s5c[key]}",f"{s6c[key]}",f"{s7c[key]}",f"{s8c[key]}"])
        Overall.append(["SPI","{:.2f}".format(s1u[key]/s1c[key]),"{:.2f}".format(s2u[key]/s2c[key]),"{:.2f}".format(s3u[key]/s3c[key]),"{:.2f}".format(s4u[key]/s4c[key]),"{:.2f}".format(s5u[key]/s5c[key]),"{:.2f}".format(s6u[key]/s6c[key]),"{:.2f}".format(s7u[key]/s7c[key]),"{:.2f}".format(s8u[key]/s8c[key])])
        Overall.append(["Total Credits Taken",f"{s1c[key]}",f"{s1c[key]+s2c[key]}",f"{s1c[key]+s2c[key]+s3c[key]}",f"{s1c[key]+s2c[key]+s3c[key]+s4c[key]}",f"{s1c[key]+s2c[key]+s3c[key]+s4c[key]+s5c[key]}",f"{s1c[key]+s2c[key]+s3c[key]+s4c[key]+s5c[key]+s6c[key]}",f"{s1c[key]+s2c[key]+s3c[key]+s4c[key]+s5c[key]+s6c[key]+s7c[key]}",f"{s1c[key]+s2c[key]+s3c[key]+s4c[key]+s5c[key]+s6c[key]+s7c[key]+s8c[key]}"])
        Overall.append(["CPI","{:.2f}".format(s1u[key]/s1c[key]),"{:.2f}".format((s1u[key]+s2u[key])/(s1c[key]+s2c[key])),"{:.2f}".format((s1u[key]+s2u[key]+s3u[key])/(s1c[key]+s2c[key]+s3c[key])),"{:.2f}".format((s1u[key]+s2u[key]+s3u[key]+s4u[key])/(s1c[key]+s2c[key]+s3c[key]+s4c[key])),"{:.2f}".format((s1u[key]+s2u[key]+s3u[key]+s4u[key]+s5u[key])/(s1c[key]+s2c[key]+s3c[key]+s4c[key]+s5c[key])),"{:.2f}".format((s1u[key]+s2u[key]+s3u[key]+s4u[key]+s5u[key]+s6u[key])/(s1c[key]+s2c[key]+s3c[key]+s4c[key]+s5c[key]+s6c[key])),"{:.2f}".format((s1u[key]+s2u[key]+s3u[key]+s4u[key]+s5u[key]+s6u[key]+s7u[key])/(s1c[key]+s2c[key]+s3c[key]+s4c[key]+s5c[key]+s6c[key]+s7c[key])),"{:.2f}".format((s1u[key]+s2u[key]+s3u[key]+s4u[key]+s5u[key]+s6u[key]+s7u[key]+s8u[key])/(s1c[key]+s2c[key]+s3c[key]+s4c[key]+s5c[key]+s6c[key]+s7c[key]+s8c[key]))])
    else: 
        Overall=wb["Overall"] 
        Overall.append(["Roll No.",f"{key}"])
        Overall.append(["Name of Student",f"{value}"])
        Overall.append(["Discipline",f"{key[4:6]}"])
        Overall.append(["Semester No.","1","2","3","4","5","6","7","8","10"])
        Overall.append(["Semester wise Credit Taken",f"{s1c[key]}",f"{s2c[key]}",f"{s3c[key]}",f"{s4c[key]}",f"{s5c[key]}",f"{s6c[key]}",f"{s7c[key]}",f"{s8c[key]}",f"{s10c[key]}"])
        Overall.append(["SPI","{:.2f}".format(s1u[key]/s1c[key]),"{:.2f}".format(s2u[key]/s2c[key]),"{:.2f}".format(s3u[key]/s3c[key]),"{:.2f}".format(s4u[key]/s4c[key]),"{:.2f}".format(s5u[key]/s5c[key]),"{:.2f}".format(s6u[key]/s6c[key]),"{:.2f}".format(s7u[key]/s7c[key]),"{:.2f}".format(s8u[key]/s8c[key]),"{:.2f}".format(s10u[key]/s10c[key])])
        Overall.append(["Total Credits Taken",f"{s1c[key]}",f"{s1c[key]+s2c[key]}",f"{s1c[key]+s2c[key]+s3c[key]}",f"{s1c[key]+s2c[key]+s3c[key]+s4c[key]}",f"{s1c[key]+s2c[key]+s3c[key]+s4c[key]+s5c[key]}",f"{s1c[key]+s2c[key]+s3c[key]+s4c[key]+s5c[key]+s6c[key]}",f"{s1c[key]+s2c[key]+s3c[key]+s4c[key]+s5c[key]+s6c[key]+s7c[key]}",f"{s1c[key]+s2c[key]+s3c[key]+s4c[key]+s5c[key]+s6c[key]+s7c[key]+s8c[key]}",f"{s1c[key]+s2c[key]+s3c[key]+s4c[key]+s5c[key]+s6c[key]+s7c[key]+s8c[key]+s10c[key]}"])
        Overall.append(["CPI","{:.2f}".format(s1u[key]/s1c[key]),"{:.2f}".format((s1u[key]+s2u[key])/(s1c[key]+s2c[key])),"{:.2f}".format((s1u[key]+s2u[key]+s3u[key])/(s1c[key]+s2c[key]+s3c[key])),"{:.2f}".format((s1u[key]+s2u[key]+s3u[key]+s4u[key])/(s1c[key]+s2c[key]+s3c[key]+s4c[key])),"{:.2f}".format((s1u[key]+s2u[key]+s3u[key]+s4u[key]+s5u[key])/(s1c[key]+s2c[key]+s3c[key]+s4c[key]+s5c[key])),"{:.2f}".format((s1u[key]+s2u[key]+s3u[key]+s4u[key]+s5u[key]+s6u[key])/(s1c[key]+s2c[key]+s3c[key]+s4c[key]+s5c[key]+s6c[key])),"{:.2f}".format((s1u[key]+s2u[key]+s3u[key]+s4u[key]+s5u[key]+s6u[key]+s7u[key])/(s1c[key]+s2c[key]+s3c[key]+s4c[key]+s5c[key]+s6c[key]+s7c[key])),"{:.2f}".format((s1u[key]+s2u[key]+s3u[key]+s4u[key]+s5u[key]+s6u[key]+s7u[key]+s8u[key])/(s1c[key]+s2c[key]+s3c[key]+s4c[key]+s5c[key]+s6c[key]+s7c[key]+s8c[key])),"{:.2f}".format((s1u[key]+s2u[key]+s3u[key]+s4u[key]+s5u[key]+s6u[key]+s7u[key]+s8u[key]+s10u[key])/(s1c[key]+s2c[key]+s3c[key]+s4c[key]+s5c[key]+s6c[key]+s7c[key]+s8c[key]+s10c[key]))])
    
    wb.save(f"output//{key}.xlsx")